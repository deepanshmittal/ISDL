import xlwings as xw
import openpyxl
import xlrd
import os

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ISDL.settings')
import django
django.setup()
from myapp.models import *

# ws = xw.Book("software final pages front sheet.xlsx").sheets['item anem']
# v1 = ws.range("G1:G5").value
# v2 = ws.range("I5").value
# print("Result:", v1, v2)
workbook = xlrd.open_workbook("data.xls")
worksheet = workbook.sheet_by_name("Sheet1")
num_cols = 19
cur_col = -1
lis = []
total_data = []
while cur_col < num_cols:
    cur_col += 1
    lis.append(worksheet.cell(0, cur_col).value)
    lis.append(worksheet.cell(1, cur_col).value)
    lis.append(worksheet.cell(2, cur_col).value)
    total_data.append(lis.copy())
    lis = []
# print(len(total_data))
# print(len(total_data[0]))
print(total_data)
for Name, ItemCode, Quantity in total_data:
    Item(Name=Name, ItemCode=ItemCode, Quantity=int(Quantity)).save()
# dataframe = openpyxl.load_workbook("data.xlsx")
# dataframe1 = dataframe.active

# for col in range(0, dataframe1.max_column):
# for row in dataframe1.iter_rows(0, dataframe1.max_row):
# print(dataframe1.iter_rows(0,dataframe1.max_row))
# print(row[col].value +"         "+row[col+1].value+"            "+row[dataframe1.max_column].value)
